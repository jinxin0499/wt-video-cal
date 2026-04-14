"""端到端集成测试：从 Excel 读取到输出报表。"""

from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

from wt_video_cal.aggregator import aggregate, apply_manager_gmv_adjustments
from wt_video_cal.calculator import calculate_all
from wt_video_cal.config import AccountInfo, AppConfig, ProfitRule
from wt_video_cal.excel_reader import read_all_excel_files
from wt_video_cal.excel_writer import write_all_reports


def _create_test_data(data_dir: Path) -> None:
    """创建3种格式的测试 Excel 文件。"""
    # 中文美区
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.append(["[日期范围]: 2026-01-01 ~ 2026-01-31"])
    ws1.append(
        [
            "达人昵称",
            "达人ID",
            "视频信息",
            "视频ID",
            "发布时间",
            "商品",
            "VV",
            "点赞数",
            "评论数",
            "分享数",
            "新增粉丝数",
            "引流次数",
            "商品曝光次数",
            "商品点击次数",
            "去重客户数",
            "视频成交订单数",
            "视频商品成交件数",
            "GMV（视频） ($)",
            "GPM ($)",
            "归因于带货视频的 GMV ($)",
        ]
    )
    ws1.append(
        [
            "digital_goodies_us",
            "7487897838011565102",
            "video info",
            "7565630769654140215",
            "2025/10/26",
            "4-in-1 Magnetic Power Bank Set – 10000mAh",
            123045,
            449,
            9,
            135,
            29,
            0,
            131635,
            4505,
            147,
            155,
            164,
            2926.18,
            23.78,
            3030.93,
        ]
    )
    wb1.save(data_dir / "shop_artcyber.xlsx")

    # 英文美区
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(["[Date Range]: 2026-01-01 ~ 2026-01-31"])
    ws2.append(
        [
            "Creator name",
            "Creator ID",
            "Video Info",
            "Video ID",
            "Time",
            "Products",
            "VV",
            "Likes",
            "Comments",
            "Shares",
            "New followers",
            "V-to-L clicks",
            "Product Impressions",
            "Product Clicks",
            "Unique customers",
            "Orders",
            "Video items sold",
            "Gross merchandise value (Video) ($)",
            "GPM ($)",
            "Shoppable video attributed GMV ($)",
        ]
    )
    ws2.append(
        [
            "user9120702396325",
            "7543951436285330487",
            "video info",
            "7592144276021087543",
            "2026/01/05",
            "Magnetic Wireless Power Bank 5000mAh",
            303,
            0,
            0,
            1,
            0,
            0,
            334,
            14,
            2,
            2,
            2,
            22.72,
            74.98,
            22.72,
        ]
    )
    wb2.save(data_dir / "shop_ac_select.xlsx")

    # 英文英区
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.append(["[Date Range]: 2026-02-01 ~ 2026-02-28"])
    ws3.append(
        [
            "Creator name",
            "Creator ID",
            "Video Info",
            "Video ID",
            "Time",
            "Products",
            "VV",
            "Likes",
            "Comments",
            "Shares",
            "New followers",
            "V-to-L clicks",
            "Product Impressions",
            "Product Clicks",
            "Unique customers",
            "Orders",
            "Video items sold",
            "Gross merchandise value (Video) (£)",
            "GPM (£)",
            "Shoppable video attributed GMV (£)",
        ]
    )
    ws3.append(
        [
            "trendyitemspot",
            "6779305119785501701",
            "video info",
            "7525893900704419074",
            "2025/07/11",
            "Upgraded 6-in-1 Magnetic Charging Kit",
            70612,
            89,
            7,
            41,
            6,
            0,
            57558,
            2695,
            54,
            57,
            60,
            1248.35,
            17.68,
            1552.14,
        ]
    )
    wb3.save(data_dir / "shop_fullera_uk.xlsx")


def _make_config() -> AppConfig:
    return AppConfig(
        default_profit_margin=Decimal("0.30"),
        profit_rules=[
            ProfitRule(
                pattern="5000",
                margin=Decimal("0.15"),
                description="5000mAh",
                max_unit_price_cny=Decimal("100"),
            ),
        ],
        accounts={
            "digital_goodies_us": AccountInfo(region="美国", manager="古嘉垚"),
            "user9120702396325": AccountInfo(region="美国", manager="刘龙海"),
            "trendyitemspot": AccountInfo(region="英国", manager="刘龙海"),
        },
        manager_monthly_gmv_usd={
            "2026-01": {
                "古嘉垚": Decimal("3100.00"),
                "刘龙海": Decimal("2050.00"),
            }
        },
    )


class TestFullPipeline:
    def test_end_to_end(self, tmp_path: Path) -> None:
        data_dir = tmp_path / "data"
        output_dir = tmp_path / "output"
        data_dir.mkdir()

        # 1. 创建测试数据
        _create_test_data(data_dir)

        # 2. 读取
        config = _make_config()
        records = read_all_excel_files(data_dir)
        assert len(records) == 3

        # 3. 计算
        results = calculate_all(
            records,
            config,
            exchange_rate_usd=Decimal("7.25"),
            exchange_rate_gbp=Decimal("9.20"),
            commission_rate=Decimal("0.05"),
        )
        assert len(results) == 3

        # 4. 汇总
        managers = aggregate(results)
        apply_manager_gmv_adjustments(
            managers,
            config,
            "2026-01",
            exchange_rate_usd=Decimal("7.25"),
            commission_rate=Decimal("0.05"),
        )
        assert len(managers) == 2  # 古嘉垚 + 刘龙海

        gu = managers["古嘉垚"]
        liu = managers["刘龙海"]

        # 古嘉垚: digital_goodies_us, GMV=3030.93 USD, 默认利润率30%
        # 3030.93 * 7.25 = 21974.24, profit = 6592.27, commission = 329.61
        assert gu.total_orders == 155
        assert gu.total_commission == Decimal("329.61")
        assert gu.gmv_diff_usd == Decimal("69.07")
        assert gu.adjustment_commission_cny == Decimal("7.51")

        # 刘龙海: user9120702396325 (USD, 5000mAh 15%) + trendyitemspot (GBP, 默认30%)
        # user9120702396325: 22.72 * 7.25 = 164.72, profit = 24.71, commission = 1.24
        # trendyitemspot: 1552.14 * 9.20 = 14279.69, profit = 4283.91, commission = 214.20
        assert liu.total_commission == Decimal("215.44")
        assert liu.gmv_diff_usd == Decimal("57.67")
        assert liu.adjustment_commission_cny == Decimal("6.27")

        # 5. 输出报表
        paths = write_all_reports(
            managers,
            output_dir,
            "2026-01",
            exchange_rate_usd=Decimal("7.25"),
            exchange_rate_gbp=Decimal("9.20"),
        )
        assert len(paths) == 3  # overview + 2 managers

        # 验证总览文件
        overview_path = output_dir / "提成汇总_2026-01.xlsx"
        assert overview_path.exists()
        wb = load_workbook(overview_path)
        ws = wb["负责人汇总"]
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 4  # header + 2 managers + total
        # 合计行
        total_row = rows[3]
        assert total_row[0] == "合计"
        assert rows[0][4] == "GMV(USD约)"
        assert rows[0][7] == "差额补算(CNY)"
        assert total_row[4] == 5023.26
        assert total_row[5] == 5150
        assert total_row[6] == 126.74
        assert total_row[7] == 13.78
        assert total_row[9] == 558.83

        ws_accounts = wb["账号明细"]
        account_rows = list(ws_accounts.iter_rows(values_only=True))
        assert len(account_rows) == 5  # header + 3 accounts + total
        assert account_rows[0][9] == "GMV(USD约)"
        assert account_rows[-1][0] == "合计"

        # 验证负责人明细文件
        gu_path = output_dir / "古嘉垚_提成明细_2026-01.xlsx"
        assert gu_path.exists()
        wb_gu = load_workbook(gu_path)
        ws_summary = wb_gu["汇总"]
        summary_rows = list(ws_summary.iter_rows(values_only=True))
        assert summary_rows[0][8] == "GMV(USD约)"
        assert summary_rows[0][12] == "差额补算(CNY)"
        assert summary_rows[-1][10] == 3100
        assert summary_rows[-1][11] == 69.07
        assert summary_rows[-1][12] == 7.51
        assert summary_rows[-1][13] == 337.12
        ws_detail = wb_gu["明细"]
        detail_rows = list(ws_detail.iter_rows(values_only=True))
        assert len(detail_rows) == 2  # header + 1 record

    def test_empty_input(self, tmp_path: Path) -> None:
        data_dir = tmp_path / "empty_data"
        data_dir.mkdir()
        records = read_all_excel_files(data_dir)
        assert records == []
