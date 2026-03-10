from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from wt_video_cal.__main__ import _check_bound_duplicates
from wt_video_cal.excel_reader import (
    detect_format,
    read_all_excel_files,
    read_excel_file,
)
from wt_video_cal.exceptions import DuplicateVideoError, UnknownFormatError
from wt_video_cal.models import Currency, ExcelFormat, VideoRecord


def _create_chinese_usd_excel(path: Path) -> None:
    """创建中文美区格式的测试 Excel。"""
    wb = Workbook()
    ws = wb.active
    ws.append(["[日期范围]: 2026-01-01 ~ 2026-01-31"])
    ws.append([
        "达人昵称", "达人ID", "视频信息", "视频ID", "发布时间", "商品",
        "VV", "点赞数", "评论数", "分享数", "新增粉丝数", "引流次数",
        "商品曝光次数", "商品点击次数", "去重客户数",
        "视频成交订单数", "视频商品成交件数",
        "GMV（视频） ($)", "GPM ($)", "归因于带货视频的 GMV ($)",
    ])
    ws.append([
        "digital_goodies_us", "7487897838011565102",
        "video info", "7565630769654140215", "2025/10/26",
        "4-in-1 Magnetic Power Bank Set",
        123045, 449, 9, 135, 29, 0, 131635, 4505, 147,
        155, 164, 2926.18, 23.78, 3030.93,
    ])
    ws.append([
        "qryzhxv", "7573313674392208439",
        "video info 2", "7595158823183174925", "2026/01/14",
        "Magnetic Wireless Power Bank 5000mAh",
        20878, 131, 2, 15, 9, 0, 21471, 972, 15,
        16, 17, 350.86, 16.81, 396.86,
    ])
    wb.save(path)


def _create_english_usd_excel(path: Path) -> None:
    """创建英文美区格式的测试 Excel。"""
    wb = Workbook()
    ws = wb.active
    ws.append(["[Date Range]: 2026-01-01 ~ 2026-01-31"])
    ws.append([
        "Creator name", "Creator ID", "Video Info", "Video ID", "Time", "Products",
        "VV", "Likes", "Comments", "Shares", "New followers", "V-to-L clicks",
        "Product Impressions", "Product Clicks", "Unique customers",
        "Orders", "Video items sold",
        "Gross merchandise value (Video) ($)", "GPM ($)",
        "Shoppable video attributed GMV ($)",
    ])
    ws.append([
        "user9120702396325", "7543951436285330487",
        "video info", "7592801730790804791", "2026/01/07",
        "Magnetic Wireless Power Bank 5000mAh",
        509, 1, 0, 0, 0, 0, 533, 9, 0, 0, 0, 0, 0, 0,
    ])
    ws.append([
        "user9120702396325", "7543951436285330487",
        "video info 2", "7592144276021087543", "2026/01/05",
        "Magnetic Wireless Power Bank 5000mAh",
        303, 0, 0, 1, 0, 0, 334, 14, 2, 2, 2, 22.72, 74.98, 22.72,
    ])
    wb.save(path)


def _create_english_gbp_excel(path: Path) -> None:
    """创建英文英区格式的测试 Excel。"""
    wb = Workbook()
    ws = wb.active
    ws.append(["[Date Range]: 2026-02-01 ~ 2026-02-28"])
    ws.append([
        "Creator name", "Creator ID", "Video Info", "Video ID", "Time", "Products",
        "VV", "Likes", "Comments", "Shares", "New followers", "V-to-L clicks",
        "Product Impressions", "Product Clicks", "Unique customers",
        "Orders", "Video items sold",
        "Gross merchandise value (Video) (£)", "GPM (£)",
        "Shoppable video attributed GMV (£)",
    ])
    ws.append([
        "trendyitemspot", "6779305119785501701",
        "video info", "7525893900704419074", "2025/07/11",
        "Upgraded 6-in-1 Magnetic Charging Kit",
        70612, 89, 7, 41, 6, 0, 57558, 2695, 54, 57, 60, 1248.35, 17.68, 1552.14,
    ])
    wb.save(path)


class TestDetectFormat:
    def test_chinese_usd(self) -> None:
        headers = ["达人昵称", "达人ID", "视频ID", "商品", "归因于带货视频的 GMV ($)"]
        assert detect_format(headers) == ExcelFormat.CHINESE_USD

    def test_english_usd(self) -> None:
        headers = [
            "Creator name", "Creator ID", "Video ID", "Products",
            "Shoppable video attributed GMV ($)",
        ]
        assert detect_format(headers) == ExcelFormat.ENGLISH_USD

    def test_english_gbp(self) -> None:
        headers = [
            "Creator name", "Creator ID", "Video ID", "Products",
            "Shoppable video attributed GMV (£)",
        ]
        assert detect_format(headers) == ExcelFormat.ENGLISH_GBP

    def test_unknown_format(self) -> None:
        with pytest.raises(UnknownFormatError):
            detect_format(["Column A", "Column B", "Column C"])


class TestReadExcelFile:
    def test_read_chinese_usd(self, tmp_path: Path) -> None:
        f = tmp_path / "chinese.xlsx"
        _create_chinese_usd_excel(f)
        records = read_excel_file(f)
        assert len(records) == 2
        assert records[0].creator_name == "digital_goodies_us"
        assert records[0].video_id == "7565630769654140215"
        assert records[0].attributed_gmv == Decimal("3030.93")
        assert records[0].currency == Currency.USD
        assert records[0].orders == 155
        assert records[0].items_sold == 164

    def test_read_english_usd(self, tmp_path: Path) -> None:
        f = tmp_path / "english_us.xlsx"
        _create_english_usd_excel(f)
        records = read_excel_file(f)
        assert len(records) == 2
        assert records[0].creator_name == "user9120702396325"
        assert records[0].currency == Currency.USD
        assert records[1].attributed_gmv == Decimal("22.72")

    def test_read_english_gbp(self, tmp_path: Path) -> None:
        f = tmp_path / "english_uk.xlsx"
        _create_english_gbp_excel(f)
        records = read_excel_file(f)
        assert len(records) == 1
        assert records[0].creator_name == "trendyitemspot"
        assert records[0].currency == Currency.GBP
        assert records[0].attributed_gmv == Decimal("1552.14")

    def test_header_with_empty_rows(self, tmp_path: Path) -> None:
        """真实 Excel: 第1行=日期, 第2行=空, 第3行=表头, 第4行起=数据。"""
        f = tmp_path / "with_gap.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["[日期范围]: 2026-01-01 ~ 2026-01-31"])
        ws.append([])  # 空行
        ws.append([
            "达人昵称", "达人ID", "视频信息", "视频ID", "发布时间", "商品",
            "VV", "点赞数", "评论数", "分享数", "新增粉丝数", "引流次数",
            "商品曝光次数", "商品点击次数", "去重客户数",
            "视频成交订单数", "视频商品成交件数",
            "GMV（视频） ($)", "GPM ($)", "归因于带货视频的 GMV ($)",
        ])
        ws.append([
            "digital_goodies_us", "7487897838011565102",
            "video info", "7565630769654140215", "2025/10/26",
            "4-in-1 Magnetic Power Bank Set",
            123045, 449, 9, 135, 29, 0, 131635, 4505, 147,
            155, 164, 2926.18, 23.78, 3030.93,
        ])
        wb.save(f)
        records = read_excel_file(f)
        assert len(records) == 1
        assert records[0].creator_name == "digital_goodies_us"
        assert records[0].attributed_gmv == Decimal("3030.93")

    def test_header_with_multiple_empty_rows(self, tmp_path: Path) -> None:
        """表头前有多个空行也能正常识别。"""
        f = tmp_path / "multi_gap.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["[Date Range]: 2026-01-01 ~ 2026-01-31"])
        ws.append([])
        ws.append([])
        ws.append([
            "Creator name", "Creator ID", "Video Info", "Video ID", "Time", "Products",
            "VV", "Likes", "Comments", "Shares", "New followers", "V-to-L clicks",
            "Product Impressions", "Product Clicks", "Unique customers",
            "Orders", "Video items sold",
            "Gross merchandise value (Video) ($)", "GPM ($)",
            "Shoppable video attributed GMV ($)",
        ])
        ws.append([
            "user9120702396325", "7543951436285330487",
            "video info", "7592144276021087543", "2026/01/05",
            "Magnetic Wireless Power Bank 5000mAh",
            303, 0, 0, 1, 0, 0, 334, 14, 2, 2, 2, 22.72, 74.98, 22.72,
        ])
        wb.save(f)
        records = read_excel_file(f)
        assert len(records) == 1
        assert records[0].creator_name == "user9120702396325"

    def test_no_header_found(self, tmp_path: Path) -> None:
        """完全没有表头的文件应抛出 UnknownFormatError。"""
        f = tmp_path / "no_header.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["random data"])
        ws.append(["more random data"])
        wb.save(f)
        with pytest.raises(UnknownFormatError):
            read_excel_file(f)

    def test_empty_file(self, tmp_path: Path) -> None:
        f = tmp_path / "empty.xlsx"
        wb = Workbook()
        wb.save(f)
        with pytest.raises(UnknownFormatError):
            read_excel_file(f)


class TestReadAllExcelFiles:
    def test_read_multiple_files(self, tmp_path: Path) -> None:
        _create_chinese_usd_excel(tmp_path / "file1.xlsx")
        _create_english_gbp_excel(tmp_path / "file2.xlsx")
        records = read_all_excel_files(tmp_path)
        assert len(records) == 3  # 2 from chinese + 1 from gbp

    def test_skip_temp_files(self, tmp_path: Path) -> None:
        _create_chinese_usd_excel(tmp_path / "data.xlsx")
        # Temp file (starts with ~$)
        _create_chinese_usd_excel(tmp_path / "~$data.xlsx")
        records = read_all_excel_files(tmp_path)
        assert len(records) == 2

    def test_empty_dir(self, tmp_path: Path) -> None:
        records = read_all_excel_files(tmp_path)
        assert records == []

    def test_nonexistent_dir(self) -> None:
        records = read_all_excel_files("/nonexistent/dir")
        assert records == []


class TestCheckDuplicates:
    def test_no_duplicates(self) -> None:
        records = [
            VideoRecord("user1", "v1", "p", Decimal("100"), 1, 1, Currency.USD, "f1"),
            VideoRecord("user1", "v2", "p", Decimal("200"), 2, 2, Currency.USD, "f1"),
        ]
        _check_bound_duplicates(records)  # should not raise

    def test_same_file_duplicates_do_not_raise(self) -> None:
        records = [
            VideoRecord("user1", "v1", "p", Decimal("100"), 1, 1, Currency.USD, "f1.xlsx"),
            VideoRecord("user1", "v1", "p", Decimal("100"), 1, 1, Currency.USD, "f1.xlsx"),
        ]
        _check_bound_duplicates(records)

    def test_cross_file_duplicates_without_orders_do_not_raise(self) -> None:
        records = [
            VideoRecord("user1", "v1", "p", Decimal("100"), 1, 1, Currency.USD, "f1.xlsx"),
            VideoRecord("user1", "v1", "p", Decimal("100"), 0, 0, Currency.USD, "f2.xlsx"),
        ]
        _check_bound_duplicates(records)

    def test_cross_file_duplicates_with_orders_raise(self) -> None:
        records = [
            VideoRecord("user1", "v1", "p", Decimal("100"), 1, 1, Currency.USD, "f1.xlsx"),
            VideoRecord("user1", "v1", "p", Decimal("100"), 2, 2, Currency.USD, "f2.xlsx"),
        ]
        with pytest.raises(DuplicateVideoError, match="跨文件重复且均有转化订单"):
            _check_bound_duplicates(records)
