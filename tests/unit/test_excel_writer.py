from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from wt_video_cal.excel_writer import write_all_reports, write_manager_report, write_overview_report
from wt_video_cal.models import (
    AccountSummary,
    CommissionResult,
    Currency,
    ManagerSummary,
    VideoRecord,
)


def _make_manager_summary() -> ManagerSummary:
    record1 = VideoRecord("acct1", "v1", "Product A", Decimal("100"), 5, 6, Currency.USD, "f")
    record2 = VideoRecord("acct2", "v2", "Product B", Decimal("200"), 10, 12, Currency.GBP, "f")

    cr1 = CommissionResult(
        record=record1, account="acct1", region="美国", manager="张三",
        exchange_rate=Decimal("7.25"), gmv_cny=Decimal("725"),
        profit_margin=Decimal("0.30"), profit_cny=Decimal("217.50"),
        commission=Decimal("10.88"),
    )
    cr2 = CommissionResult(
        record=record2, account="acct2", region="英国", manager="张三",
        exchange_rate=Decimal("9.20"), gmv_cny=Decimal("1840"),
        profit_margin=Decimal("0.30"), profit_cny=Decimal("552"),
        commission=Decimal("27.60"),
    )

    acct1 = AccountSummary(
        account="acct1", region="美国", manager="张三",
        total_orders=5, total_items_sold=6,
        total_gmv_cny=Decimal("725"), total_profit_cny=Decimal("217.50"),
        total_commission=Decimal("10.88"), details=[cr1],
    )
    acct2 = AccountSummary(
        account="acct2", region="英国", manager="张三",
        total_orders=10, total_items_sold=12,
        total_gmv_cny=Decimal("1840"), total_profit_cny=Decimal("552"),
        total_commission=Decimal("27.60"), details=[cr2],
    )

    return ManagerSummary(
        manager="张三",
        total_orders=15, total_items_sold=18,
        total_gmv_cny=Decimal("2565"), total_profit_cny=Decimal("769.50"),
        total_commission=Decimal("38.48"),
        accounts={"acct1": acct1, "acct2": acct2},
    )


class TestWriteManagerReport:
    def test_creates_file(self, tmp_path: Path) -> None:
        ms = _make_manager_summary()
        path = write_manager_report(ms, tmp_path, "2026-01")
        assert path.exists()
        assert path.name == "张三_提成明细_2026-01.xlsx"

    def test_summary_sheet(self, tmp_path: Path) -> None:
        ms = _make_manager_summary()
        path = write_manager_report(ms, tmp_path, "2026-01")
        wb = load_workbook(path)
        ws = wb["汇总"]
        rows = list(ws.iter_rows(values_only=True))
        # Header + 2 accounts + total = 4 rows
        assert len(rows) == 4
        # Check header
        assert rows[0][0] == "账号"
        # Check total row
        assert rows[3][0] == "合计"
        assert rows[3][2] == 15  # total orders

    def test_detail_sheet(self, tmp_path: Path) -> None:
        ms = _make_manager_summary()
        path = write_manager_report(ms, tmp_path, "2026-01")
        wb = load_workbook(path)
        ws = wb["明细"]
        rows = list(ws.iter_rows(values_only=True))
        # Header + 2 details
        assert len(rows) == 3
        assert rows[0][0] == "账号"
        assert rows[1][1] == "v1"


class TestWriteOverviewReport:
    def test_creates_file(self, tmp_path: Path) -> None:
        ms = _make_manager_summary()
        managers = {"张三": ms}
        path = write_overview_report(managers, tmp_path, "2026-01")
        assert path.exists()
        assert path.name == "提成汇总_2026-01.xlsx"

    def test_manager_summary_sheet(self, tmp_path: Path) -> None:
        ms = _make_manager_summary()
        managers = {"张三": ms}
        path = write_overview_report(managers, tmp_path, "2026-01")
        wb = load_workbook(path)
        ws = wb["负责人汇总"]
        rows = list(ws.iter_rows(values_only=True))
        # Header + 1 manager + total = 3
        assert len(rows) == 3
        assert rows[1][0] == "张三"
        assert rows[2][0] == "合计"

    def test_account_detail_sheet(self, tmp_path: Path) -> None:
        ms = _make_manager_summary()
        managers = {"张三": ms}
        path = write_overview_report(managers, tmp_path, "2026-01")
        wb = load_workbook(path)
        ws = wb["账号明细"]
        rows = list(ws.iter_rows(values_only=True))
        # Header + 2 accounts
        assert len(rows) == 3


    def test_analysis_sheet_exists(self, tmp_path: Path) -> None:
        ms = _make_manager_summary()
        path = write_manager_report(ms, tmp_path, "2026-01")
        wb = load_workbook(path)
        assert "分析" in wb.sheetnames

    def test_analysis_sheet_no_manager_column(self, tmp_path: Path) -> None:
        ms = _make_manager_summary()
        path = write_manager_report(ms, tmp_path, "2026-01")
        wb = load_workbook(path)
        ws = wb["分析"]
        # Find "账号业绩排名" section and check it has no "负责人" column
        for row in ws.iter_rows(values_only=True):
            if row[0] == "账号业绩排名":
                # Next row is the header
                break
        # Read the header row after the section title
        found_section = False
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            if values[0] == "账号业绩排名":
                found_section = True
                continue
            if found_section:
                # This is the header row
                assert "负责人" not in values
                break

    def test_analysis_section_titles(self, tmp_path: Path) -> None:
        ms = _make_manager_summary()
        path = write_manager_report(ms, tmp_path, "2026-01")
        wb = load_workbook(path)
        ws = wb["分析"]
        section_titles = set()
        for row in ws.iter_rows(values_only=True):
            val = row[0]
            if isinstance(val, str) and val in {
                "Top 10 视频 — 按订单数", "Top 10 视频 — 按GMV(CNY)",
                "Top 10 商品 — 按销量", "Top 10 商品 — 按GMV(CNY)",
                "账号业绩排名", "区域分布", "利润率分布",
            }:
                section_titles.add(val)
        assert len(section_titles) == 7


class TestWriteOverviewAnalysis:
    def test_analysis_sheet_exists(self, tmp_path: Path) -> None:
        ms = _make_manager_summary()
        managers = {"张三": ms}
        path = write_overview_report(managers, tmp_path, "2026-01")
        wb = load_workbook(path)
        assert "分析" in wb.sheetnames

    def test_analysis_sheet_has_manager_column(self, tmp_path: Path) -> None:
        ms = _make_manager_summary()
        managers = {"张三": ms}
        path = write_overview_report(managers, tmp_path, "2026-01")
        wb = load_workbook(path)
        ws = wb["分析"]
        found_section = False
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            if values[0] == "账号业绩排名":
                found_section = True
                continue
            if found_section:
                assert "负责人" in values
                break


class TestWriteAllReports:
    def test_generates_all_files(self, tmp_path: Path) -> None:
        ms = _make_manager_summary()
        managers = {"张三": ms}
        paths = write_all_reports(managers, tmp_path, "2026-01")
        assert len(paths) == 2  # overview + 1 manager
        names = {p.name for p in paths}
        assert "提成汇总_2026-01.xlsx" in names
        assert "张三_提成明细_2026-01.xlsx" in names
