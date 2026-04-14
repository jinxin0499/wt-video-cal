import logging
from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from wt_video_cal.analysis import (
    AnalysisResult,
    compute_analysis,
    extract_all_details,
    extract_manager_details,
)
from wt_video_cal.models import (
    AccountSummary,
    LowMarginReviewItem,
    ManagerSummary,
    SourceCreatorStats,
)
from wt_video_cal.settings import (
    EXCHANGE_RATE_GBP,
    EXCHANGE_RATE_JPY,
    EXCHANGE_RATE_USD,
    REPORT_MONTH,
)

logger = logging.getLogger(__name__)

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_TOTAL_FONT = Font(bold=True, color="C00000")
_TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_ALIGN_RIGHT = Alignment(horizontal="right")
_ALIGN_CENTER = Alignment(horizontal="center")
_SECTION_FONT = Font(bold=True, size=12)
_SECTION_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")


def _fmt(value: Decimal) -> float:
    """Decimal → float for Excel cell (保留2位精度)。"""
    return float(value.quantize(Decimal("0.01")))


def _fmt4(value: Decimal) -> float:
    """Decimal → float for Excel cell (保留4位精度，用于汇率)。"""
    return float(value.quantize(Decimal("0.0001")))


def _fmt_optional(value: Decimal | None) -> float | str:
    if value is None:
        return ""
    return _fmt(value)


def _cny_to_usd(value_cny: Decimal, exchange_rate_usd: Decimal = EXCHANGE_RATE_USD) -> Decimal:
    if exchange_rate_usd == Decimal("0"):
        return Decimal("0")
    return value_cny / exchange_rate_usd


def _original_gmv_to_cny(
    gmv: Decimal,
    currency: str,
    *,
    exchange_rate_usd: Decimal = EXCHANGE_RATE_USD,
    exchange_rate_gbp: Decimal = EXCHANGE_RATE_GBP,
    exchange_rate_jpy: Decimal = EXCHANGE_RATE_JPY,
) -> Decimal:
    if currency == "GBP":
        return gmv * exchange_rate_gbp
    if currency == "JPY":
        return gmv * exchange_rate_jpy
    return gmv * exchange_rate_usd


def _original_gmv_to_usd(
    gmv: Decimal,
    currency: str,
    *,
    exchange_rate_usd: Decimal = EXCHANGE_RATE_USD,
    exchange_rate_gbp: Decimal = EXCHANGE_RATE_GBP,
    exchange_rate_jpy: Decimal = EXCHANGE_RATE_JPY,
) -> Decimal:
    gmv_cny = _original_gmv_to_cny(
        gmv,
        currency,
        exchange_rate_usd=exchange_rate_usd,
        exchange_rate_gbp=exchange_rate_gbp,
        exchange_rate_jpy=exchange_rate_jpy,
    )
    return _cny_to_usd(gmv_cny, exchange_rate_usd)


def _style_header(ws: object, row: int, col_count: int) -> None:
    """为表头行设置样式。"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)  # type: ignore[union-attr]
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _ALIGN_CENTER


def _auto_width(ws: object, col_count: int) -> None:
    """自动调整列宽。"""
    for col in range(1, col_count + 1):
        max_len = 0
        letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):  # type: ignore[union-attr]
            for cell in row:
                if cell.value is not None:
                    # 中文字符按2倍宽度
                    val_str = str(cell.value)
                    char_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                    max_len = max(max_len, char_len)
        ws.column_dimensions[letter].width = min(max_len + 4, 50)  # type: ignore[union-attr]


def _write_summary_row(ws: object, row: int, col_count: int, values: Sequence[object]) -> None:
    """写入合计行并设置样式。"""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col)  # type: ignore[union-attr]
        cell.value = val
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_FILL


def _get_account_currency_info(
    acct: AccountSummary,
) -> tuple[str, Decimal, Decimal]:
    """从明细记录中提取账号的币种、汇率、原币种GMV合计。

    若账号含多币种记录，以GMV较大的币种为主，币种标记加"*"。
    """
    currency_gmv: dict[str, Decimal] = defaultdict(Decimal)
    currency_rate: dict[str, Decimal] = {}
    for d in acct.details:
        cur = d.record.currency.value
        currency_gmv[cur] += d.record.attributed_gmv
        currency_rate[cur] = d.exchange_rate

    if len(currency_gmv) == 0:
        return "", Decimal("0"), Decimal("0")

    if len(currency_gmv) == 1:
        cur = next(iter(currency_gmv))
        return cur, currency_rate[cur], currency_gmv[cur]

    # 多币种：按GMV金额取主币种，标记"*"
    primary = max(currency_gmv, key=lambda c: currency_gmv[c])
    return primary + "*", currency_rate[primary], sum(currency_gmv.values(), Decimal("0"))


def _write_section_title(ws: object, row: int, title: str, col_span: int) -> None:
    """写入合并单元格的区域标题。"""
    cell = ws.cell(row=row, column=1)  # type: ignore[union-attr]
    cell.value = title
    cell.font = _SECTION_FONT
    cell.fill = _SECTION_FILL
    cell.alignment = Alignment(horizontal="left")
    if col_span > 1:
        ws.merge_cells(  # type: ignore[union-attr]
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=col_span,
        )


def _write_analysis_table(
    ws: object,
    start_row: int,
    section_title: str,
    headers: list[str],
    rows: list[list[object]],
) -> int:
    """写入一个分析表（标题+表头+数据行），返回下一个可用行号。"""
    col_count = len(headers)

    # 区域标题
    _write_section_title(ws, start_row, section_title, col_count)
    start_row += 1

    # 表头
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col)  # type: ignore[union-attr]
        cell.value = h
    _style_header(ws, start_row, col_count)
    start_row += 1

    # 数据行
    for row_data in rows:
        for col, val in enumerate(row_data, 1):
            ws.cell(row=start_row, column=col).value = val  # type: ignore[union-attr]
        start_row += 1

    # 空行间隔
    return start_row + 1


def _write_analysis_sheet(
    ws: object,
    analysis: AnalysisResult,
    *,
    include_manager_column: bool = True,
) -> None:
    """写入完整的分析 sheet。"""
    row = 1

    # --- 1. Top 10 视频 — 按订单数 ---
    video_headers = ["排名", "视频ID", "商品", "账号", "订单数", "成交件数", "GMV(CNY)"]
    video_rows_orders = [
        [v.rank, v.video_id, v.product_name, v.account, v.orders, v.items_sold, _fmt(v.gmv_cny)]
        for v in analysis.top_videos_by_orders
    ]
    row = _write_analysis_table(ws, row, "Top 10 视频 — 按订单数", video_headers, video_rows_orders)

    # --- 2. Top 10 视频 — 按GMV(CNY) ---
    video_rows_gmv = [
        [v.rank, v.video_id, v.product_name, v.account, v.orders, v.items_sold, _fmt(v.gmv_cny)]
        for v in analysis.top_videos_by_gmv
    ]
    row = _write_analysis_table(ws, row, "Top 10 视频 — 按GMV(CNY)", video_headers, video_rows_gmv)

    # --- 3. Top 10 商品 — 按销量 ---
    product_headers = ["排名", "商品", "订单数", "成交件数", "GMV(CNY)", "占比(件数)", "占比(GMV)"]
    product_rows_items = [
        [
            p.rank,
            p.product_name,
            p.orders,
            p.items_sold,
            _fmt(p.gmv_cny),
            f"{_fmt(p.items_pct)}%",
            f"{_fmt(p.gmv_pct)}%",
        ]
        for p in analysis.top_products_by_items
    ]
    row = _write_analysis_table(
        ws,
        row,
        "Top 10 商品 — 按销量",
        product_headers,
        product_rows_items,
    )

    # --- 4. Top 10 商品 — 按GMV(CNY) ---
    product_rows_gmv = [
        [
            p.rank,
            p.product_name,
            p.orders,
            p.items_sold,
            _fmt(p.gmv_cny),
            f"{_fmt(p.items_pct)}%",
            f"{_fmt(p.gmv_pct)}%",
        ]
        for p in analysis.top_products_by_gmv
    ]
    row = _write_analysis_table(
        ws,
        row,
        "Top 10 商品 — 按GMV(CNY)",
        product_headers,
        product_rows_gmv,
    )

    # --- 5. 账号业绩排名 ---
    if include_manager_column:
        acct_headers = [
            "排名",
            "账号",
            "负责人",
            "区域",
            "订单数",
            "成交件数",
            "GMV(CNY)",
            "提成(CNY)",
            "客单价(CNY)",
        ]
        acct_rows = [
            [
                a.rank,
                a.account,
                a.manager,
                a.region,
                a.orders,
                a.items_sold,
                _fmt(a.gmv_cny),
                _fmt(a.commission),
                _fmt(a.unit_price),
            ]
            for a in analysis.account_rankings
        ]
    else:
        acct_headers = [
            "排名",
            "账号",
            "区域",
            "订单数",
            "成交件数",
            "GMV(CNY)",
            "提成(CNY)",
            "客单价(CNY)",
        ]
        acct_rows = [
            [
                a.rank,
                a.account,
                a.region,
                a.orders,
                a.items_sold,
                _fmt(a.gmv_cny),
                _fmt(a.commission),
                _fmt(a.unit_price),
            ]
            for a in analysis.account_rankings
        ]
    row = _write_analysis_table(ws, row, "账号业绩排名", acct_headers, acct_rows)

    # --- 6. 区域分布 ---
    region_headers = ["区域", "账号数", "订单数", "成交件数", "GMV(CNY)", "提成(CNY)", "GMV占比"]
    region_rows = [
        [
            r.region,
            r.account_count,
            r.orders,
            r.items_sold,
            _fmt(r.gmv_cny),
            _fmt(r.commission),
            f"{_fmt(r.gmv_pct)}%",
        ]
        for r in analysis.region_breakdown
    ]
    row = _write_analysis_table(ws, row, "区域分布", region_headers, region_rows)

    # --- 7. 利润率分布 ---
    margin_headers = ["利润率", "商品数", "订单数", "成交件数", "GMV(CNY)", "GMV占比"]
    margin_rows = [
        [
            m.margin,
            m.product_count,
            m.orders,
            m.items_sold,
            _fmt(m.gmv_cny),
            f"{_fmt(m.gmv_pct)}%",
        ]
        for m in analysis.margin_distribution
    ]
    _write_analysis_table(ws, row, "利润率分布", margin_headers, margin_rows)

    # 自动列宽 — 使用最大列数
    max_cols = max(
        len(video_headers),
        len(product_headers),
        len(acct_headers),
        len(region_headers),
        len(margin_headers),
    )
    _auto_width(ws, max_cols)


def write_manager_report(
    manager_summary: ManagerSummary,
    output_dir: Path,
    report_month: str = REPORT_MONTH,
    exchange_rate_usd: Decimal = EXCHANGE_RATE_USD,
) -> Path:
    """为单位负责人生成提成明细报表。"""
    wb = Workbook()

    # === Sheet 1: 汇总 ===
    ws_summary = wb.active
    if ws_summary is None:
        raise RuntimeError("Workbook 缺少活动工作表")
    ws_summary.title = "汇总"

    summary_headers = [
        "账号",
        "区域",
        "订单数",
        "成交件数",
        "GMV(原币种)",
        "币种",
        "汇率",
        "GMV(CNY)",
        "GMV(USD约)",
        "提成(CNY)",
        "达人端GMV(USD)",
        "GMV差额(USD)",
        "差额补算(CNY)",
        "提成合计(CNY)",
    ]
    ws_summary.append(summary_headers)  # type: ignore[union-attr]
    _style_header(ws_summary, 1, len(summary_headers))

    row_num = 2
    for acct in sorted(manager_summary.accounts.values(), key=lambda a: a.account):
        currency, exchange_rate, gmv_original = _get_account_currency_info(acct)
        ws_summary.append(
            [  # type: ignore[union-attr]
                acct.account,
                acct.region,
                acct.total_orders,
                acct.total_items_sold,
                _fmt(gmv_original),
                currency,
                _fmt4(exchange_rate),
                _fmt(acct.total_gmv_cny),
                _fmt(_cny_to_usd(acct.total_gmv_cny, exchange_rate_usd)),
                _fmt(acct.total_commission),
                "",
                "",
                "",
                "",
            ]
        )
        row_num += 1

    # 合计行
    _write_summary_row(
        ws_summary,
        row_num,
        len(summary_headers),
        [
            "合计",
            "",
            manager_summary.total_orders,
            manager_summary.total_items_sold,
            "",
            "",
            "",
            _fmt(manager_summary.total_gmv_cny),
            _fmt(_cny_to_usd(manager_summary.total_gmv_cny, exchange_rate_usd)),
            _fmt(manager_summary.total_commission),
            _fmt_optional(manager_summary.creator_side_gmv_usd),
            _fmt_optional(manager_summary.gmv_diff_usd),
            _fmt(manager_summary.adjustment_commission_cny),
            _fmt(manager_summary.total_commission_with_adjustment),
        ],
    )

    _auto_width(ws_summary, len(summary_headers))

    # === Sheet 2: 明细 ===
    ws_detail = wb.create_sheet("明细")
    detail_headers = [
        "账号",
        "视频ID",
        "商品",
        "订单数",
        "商品件数",
        "GMV(原币种)",
        "币种",
        "汇率",
        "GMV(CNY)",
        "利润率",
        "提成(CNY)",
    ]
    ws_detail.append(detail_headers)
    _style_header(ws_detail, 1, len(detail_headers))

    for acct in sorted(manager_summary.accounts.values(), key=lambda a: a.account):
        for detail in acct.details:
            ws_detail.append(
                [
                    detail.account,
                    detail.record.video_id,
                    detail.record.product_name,
                    detail.record.orders,
                    detail.record.items_sold,
                    _fmt(detail.record.attributed_gmv),
                    detail.record.currency.value,
                    _fmt4(detail.exchange_rate),
                    _fmt(detail.gmv_cny),
                    f"{_fmt(detail.profit_margin * 100)}%",
                    _fmt(detail.commission),
                ]
            )

    _auto_width(ws_detail, len(detail_headers))

    # === Sheet 3: 分析 ===
    ws_analysis = wb.create_sheet("分析")
    mgr_details = extract_manager_details(manager_summary)
    analysis = compute_analysis(mgr_details)
    _write_analysis_sheet(ws_analysis, analysis, include_manager_column=False)

    # 保存
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{manager_summary.manager}_提成明细_{report_month}.xlsx"
    output_path = output_dir / filename
    wb.save(output_path)
    logger.info("已生成: %s", output_path)
    return output_path


def write_overview_report(
    managers: dict[str, ManagerSummary],
    output_dir: Path,
    report_month: str = REPORT_MONTH,
    source_summary: dict[str, dict[str, SourceCreatorStats]] | None = None,
    low_margin_review_items: list[LowMarginReviewItem] | None = None,
    exchange_rate_usd: Decimal = EXCHANGE_RATE_USD,
    exchange_rate_gbp: Decimal = EXCHANGE_RATE_GBP,
    exchange_rate_jpy: Decimal = EXCHANGE_RATE_JPY,
) -> Path:
    """生成总览报表。"""
    wb = Workbook()

    # === Sheet 1: 负责人汇总 ===
    ws_managers = wb.active
    if ws_managers is None:
        raise RuntimeError("Workbook 缺少活动工作表")
    ws_managers.title = "负责人汇总"

    mgr_headers = [
        "负责人",
        "订单数",
        "成交件数",
        "GMV(CNY)",
        "GMV(USD约)",
        "达人端GMV(USD)",
        "GMV差额(USD)",
        "差额补算(CNY)",
        "提成(CNY)",
        "提成合计(CNY)",
    ]
    ws_managers.append(mgr_headers)  # type: ignore[union-attr]
    _style_header(ws_managers, 1, len(mgr_headers))

    total_orders = 0
    total_items = 0
    total_gmv = Decimal("0")
    total_creator_side_gmv_usd = Decimal("0")
    total_gmv_diff_usd = Decimal("0")
    total_adjustment_commission = Decimal("0")
    total_commission = Decimal("0")

    row_num = 2
    for ms in sorted(managers.values(), key=lambda m: m.manager):
        ws_managers.append(
            [  # type: ignore[union-attr]
                ms.manager,
                ms.total_orders,
                ms.total_items_sold,
                _fmt(ms.total_gmv_cny),
                _fmt(_cny_to_usd(ms.total_gmv_cny, exchange_rate_usd)),
                _fmt_optional(ms.creator_side_gmv_usd),
                _fmt_optional(ms.gmv_diff_usd),
                _fmt(ms.adjustment_commission_cny),
                _fmt(ms.total_commission),
                _fmt(ms.total_commission_with_adjustment),
            ]
        )
        total_orders += ms.total_orders
        total_items += ms.total_items_sold
        total_gmv += ms.total_gmv_cny
        if ms.creator_side_gmv_usd is not None:
            total_creator_side_gmv_usd += ms.creator_side_gmv_usd
        if ms.gmv_diff_usd is not None:
            total_gmv_diff_usd += ms.gmv_diff_usd
        total_adjustment_commission += ms.adjustment_commission_cny
        total_commission += ms.total_commission
        row_num += 1

    _write_summary_row(
        ws_managers,
        row_num,
        len(mgr_headers),
        [
            "合计",
            total_orders,
            total_items,
            _fmt(total_gmv),
            _fmt(_cny_to_usd(total_gmv, exchange_rate_usd)),
            _fmt(total_creator_side_gmv_usd),
            _fmt(total_gmv_diff_usd),
            _fmt(total_adjustment_commission),
            _fmt(total_commission),
            _fmt(total_commission + total_adjustment_commission),
        ],
    )

    _auto_width(ws_managers, len(mgr_headers))

    # === Sheet 2: 账号明细 ===
    ws_accounts = wb.create_sheet("账号明细")
    acct_headers = [
        "负责人",
        "账号",
        "区域",
        "订单数",
        "成交件数",
        "GMV(原币种)",
        "币种",
        "汇率",
        "GMV(CNY)",
        "GMV(USD约)",
        "提成(CNY)",
    ]
    ws_accounts.append(acct_headers)
    _style_header(ws_accounts, 1, len(acct_headers))

    for ms in sorted(managers.values(), key=lambda m: m.manager):
        for acct in sorted(ms.accounts.values(), key=lambda a: a.account):
            currency, exchange_rate, gmv_original = _get_account_currency_info(acct)
            ws_accounts.append(
                [
                    ms.manager,
                    acct.account,
                    acct.region,
                    acct.total_orders,
                    acct.total_items_sold,
                    _fmt(gmv_original),
                    currency,
                    _fmt4(exchange_rate),
                    _fmt(acct.total_gmv_cny),
                    _fmt(_cny_to_usd(acct.total_gmv_cny, exchange_rate_usd)),
                    _fmt(acct.total_commission),
                ]
            )

    row_num = ws_accounts.max_row + 1
    _write_summary_row(
        ws_accounts,
        row_num,
        len(acct_headers),
        [
            "合计",
            "",
            "",
            total_orders,
            total_items,
            "",
            "",
            "",
            _fmt(total_gmv),
            _fmt(_cny_to_usd(total_gmv, exchange_rate_usd)),
            _fmt(total_commission),
        ],
    )

    _auto_width(ws_accounts, len(acct_headers))

    # === Sheet 3: 数据校验（按来源文件汇总绑定账号原始数据） ===
    if source_summary:
        ws_verify = wb.create_sheet("数据校验")
        verify_headers = ["来源文件", "达人", "订单数", "成交件数", "GMV", "币种", "GMV(USD约)"]
        ws_verify.append(verify_headers)
        _style_header(ws_verify, 1, len(verify_headers))

        grand_orders = 0
        grand_items = 0
        grand_gmv = Decimal("0")
        grand_gmv_usd = Decimal("0")

        for fname in sorted(source_summary):
            creators = source_summary[fname]
            for creator in sorted(creators):
                s = creators[creator]
                ws_verify.append(
                    [
                        fname,
                        creator,
                        s["orders"],
                        s["items_sold"],
                        _fmt(s["gmv"]),
                        s["currency"],
                        _fmt(
                            _original_gmv_to_usd(
                                s["gmv"],
                                s["currency"],
                                exchange_rate_usd=exchange_rate_usd,
                                exchange_rate_gbp=exchange_rate_gbp,
                                exchange_rate_jpy=exchange_rate_jpy,
                            )
                        ),
                    ]
                )

            # 文件小计
            file_orders = sum(s["orders"] for s in creators.values())
            file_items = sum(s["items_sold"] for s in creators.values())
            file_gmv = sum((s["gmv"] for s in creators.values()), Decimal("0"))
            file_gmv_usd = sum(
                (
                    _original_gmv_to_usd(
                        s["gmv"],
                        s["currency"],
                        exchange_rate_usd=exchange_rate_usd,
                        exchange_rate_gbp=exchange_rate_gbp,
                        exchange_rate_jpy=exchange_rate_jpy,
                    )
                    for s in creators.values()
                ),
                Decimal("0"),
            )
            row_num = ws_verify.max_row + 1
            _write_summary_row(
                ws_verify,
                row_num,
                len(verify_headers),
                [
                    f"小计: {fname}",
                    "",
                    file_orders,
                    file_items,
                    _fmt(file_gmv),
                    "",
                    _fmt(file_gmv_usd),
                ],
            )
            grand_orders += file_orders
            grand_items += file_items
            grand_gmv += file_gmv
            grand_gmv_usd += file_gmv_usd

        # 总合计
        row_num = ws_verify.max_row + 1
        _write_summary_row(
            ws_verify,
            row_num,
            len(verify_headers),
            [
                "合计",
                "",
                grand_orders,
                grand_items,
                _fmt(grand_gmv),
                "",
                _fmt(grand_gmv_usd),
            ],
        )

        _auto_width(ws_verify, len(verify_headers))

    # === Sheet 4: 低毛利复核 ===
    if low_margin_review_items:
        ws_review = wb.create_sheet("低毛利复核")
        review_headers = [
            "来源文件",
            "负责人",
            "账号",
            "区域",
            "视频ID",
            "商品",
            "币种",
            "GMV(视频)",
            "订单数",
            "成交件数",
            "件单价(原币种)",
            "件单价(CNY)",
            "命中关键词",
            "规则利润率",
            "阈值(CNY)",
            "未按低毛利原因",
        ]
        ws_review.append(review_headers)
        _style_header(ws_review, 1, len(review_headers))

        for item in low_margin_review_items:
            ws_review.append(
                [
                    item.source_file,
                    item.manager,
                    item.account,
                    item.region,
                    item.video_id,
                    item.product_name,
                    item.currency,
                    _fmt(item.video_gmv),
                    item.orders,
                    item.items_sold,
                    _fmt_optional(item.unit_price_original),
                    _fmt_optional(item.unit_price_cny),
                    item.matched_pattern,
                    f"{_fmt(item.rule_margin * 100)}%",
                    _fmt_optional(item.max_unit_price_cny),
                    item.reason,
                ]
            )

        _auto_width(ws_review, len(review_headers))

    # === Sheet: 分析 ===
    ws_analysis = wb.create_sheet("分析")
    all_details = extract_all_details(managers)
    analysis = compute_analysis(all_details)
    _write_analysis_sheet(ws_analysis, analysis, include_manager_column=True)

    # 保存
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"提成汇总_{report_month}.xlsx"
    output_path = output_dir / filename
    wb.save(output_path)
    logger.info("已生成: %s", output_path)
    return output_path


def write_all_reports(
    managers: dict[str, ManagerSummary],
    output_dir: str | Path,
    report_month: str = REPORT_MONTH,
    source_summary: dict[str, dict[str, SourceCreatorStats]] | None = None,
    low_margin_review_items: list[LowMarginReviewItem] | None = None,
    exchange_rate_usd: Decimal = EXCHANGE_RATE_USD,
    exchange_rate_gbp: Decimal = EXCHANGE_RATE_GBP,
    exchange_rate_jpy: Decimal = EXCHANGE_RATE_JPY,
) -> list[Path]:
    """生成所有报表：总览 + 每位负责人明细。"""
    out = Path(output_dir)
    paths: list[Path] = []

    # 总览
    paths.append(
        write_overview_report(
            managers,
            out,
            report_month,
            source_summary,
            low_margin_review_items,
            exchange_rate_usd,
            exchange_rate_gbp,
            exchange_rate_jpy,
        )
    )

    # 每位负责人
    for ms in managers.values():
        paths.append(write_manager_report(ms, out, report_month, exchange_rate_usd))

    return paths
