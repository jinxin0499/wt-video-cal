import logging
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import cast

from openpyxl import load_workbook

from wt_video_cal.exceptions import UnknownFormatError
from wt_video_cal.models import Currency, ExcelFormat, VideoRecord

logger = logging.getLogger(__name__)

# 用于识别表头行的关键词
_HEADER_MARKERS = {"达人昵称", "Creator name"}


def _matches_header(actual: str, expected: str | tuple[str, ...]) -> bool:
    if isinstance(expected, tuple):
        return actual in expected
    return actual == expected


def _header_label(expected: str | tuple[str, ...]) -> str:
    if isinstance(expected, tuple):
        return " / ".join(expected)
    return expected


# 各格式的表头关键字段映射
_FORMAT_COLUMNS: dict[ExcelFormat, dict[str, str | tuple[str, ...]]] = {
    ExcelFormat.CHINESE_USD: {
        "creator_name": "达人昵称",
        "video_id": "视频ID",
        "product_name": "商品",
        "video_gmv": "GMV（视频） ($)",
        "attributed_gmv": "归因于带货视频的 GMV ($)",
        "orders": "视频成交订单数",
        "items_sold": "视频商品成交件数",
    },
    ExcelFormat.CHINESE_JPY: {
        "creator_name": "达人昵称",
        "video_id": "视频ID",
        "product_name": "商品",
        "video_gmv": (
            "GMV（视频） (円)",
            "GMV（视频） (¥)",
            "GMV（视频） (￥)",
            "GMV（视频） (JPY)",
        ),
        "attributed_gmv": (
            "归因于带货视频的 GMV (円)",
            "归因于带货视频的 GMV (¥)",
            "归因于带货视频的 GMV (￥)",
            "归因于带货视频的 GMV (JPY)",
        ),
        "orders": "视频成交订单数",
        "items_sold": "视频商品成交件数",
    },
    ExcelFormat.ENGLISH_USD: {
        "creator_name": "Creator name",
        "video_id": "Video ID",
        "product_name": "Products",
        "video_gmv": "Gross merchandise value (Video) ($)",
        "attributed_gmv": "Shoppable video attributed GMV ($)",
        "orders": "Orders",
        "items_sold": "Video items sold",
    },
    ExcelFormat.ENGLISH_GBP: {
        "creator_name": "Creator name",
        "video_id": "Video ID",
        "product_name": "Products",
        "video_gmv": "Gross merchandise value (Video) (£)",
        "attributed_gmv": "Shoppable video attributed GMV (£)",
        "orders": "Orders",
        "items_sold": "Video items sold",
    },
    ExcelFormat.ENGLISH_JPY: {
        "creator_name": "Creator name",
        "video_id": "Video ID",
        "product_name": "Products",
        "video_gmv": (
            "Gross merchandise value (Video) (¥)",
            "Gross merchandise value (Video) (￥)",
            "Gross merchandise value (Video) (JPY)",
        ),
        "attributed_gmv": (
            "Shoppable video attributed GMV (¥)",
            "Shoppable video attributed GMV (￥)",
            "Shoppable video attributed GMV (JPY)",
        ),
        "orders": "Orders",
        "items_sold": "Video items sold",
    },
}


def _is_header_row(row: tuple[object, ...]) -> bool:
    """判断一行是否为表头行。"""
    for cell in row:
        if cell is not None and str(cell).strip() in _HEADER_MARKERS:
            return True
    return False


def _is_empty_export(rows: list[tuple[object, ...]]) -> bool:
    """识别仅含日期范围、没有表头和数据的空导出。"""
    non_empty_rows = []
    for row in rows:
        values = [str(cell).strip() for cell in row if cell not in (None, "")]
        if values:
            non_empty_rows.append(values)

    if len(non_empty_rows) != 1:
        return False

    first_cell = non_empty_rows[0][0]
    return first_cell.startswith("[日期范围]:") or first_cell.startswith("[Date Range]:")


def _load_rows(file_path: Path) -> list[tuple[object, ...]]:
    """加载 Excel 行数据。优先 read_only 模式，若结果异常则回退到普通模式。"""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []
    rows = cast(list[tuple[object, ...]], list(ws.iter_rows(values_only=True)))
    wb.close()

    # 某些 TikTok 导出文件使用 inline string (t="str")，read_only 模式无法正确解析。
    # 实际表现通常是只读到日期范围等极少行，因此当行数异常少时直接回退普通模式重读。
    if len(rows) <= 5:
        logger.debug(
            "[%s] read_only 模式仅读到 %d 行，回退到普通模式",
            file_path.name,
            len(rows),
        )
        wb = load_workbook(file_path, read_only=False, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return []
        rows = cast(list[tuple[object, ...]], list(ws.iter_rows(values_only=True)))
        wb.close()
        logger.debug("[%s] 普通模式读到 %d 行", file_path.name, len(rows))

    return rows


def detect_format(header_row: list[str]) -> ExcelFormat:
    """根据表头行自动识别 Excel 格式。"""
    headers_str = "\t".join(header_row)

    if "达人昵称" in headers_str:
        for h in header_row:
            if "GMV (円)" in h or "GMV（视频） (円)" in h or "GMV (JPY)" in h:
                return ExcelFormat.CHINESE_JPY
            if "GMV (¥)" in h or "GMV (￥)" in h:
                return ExcelFormat.CHINESE_JPY
        return ExcelFormat.CHINESE_USD

    if "Creator name" in headers_str:
        # 区分美区、英区、日区
        for h in header_row:
            if "GMV (£)" in h or "GMV (£)" in h:
                return ExcelFormat.ENGLISH_GBP
            if "GMV (¥)" in h or "GMV (￥)" in h or "GMV (JPY)" in h:
                return ExcelFormat.ENGLISH_JPY
            if "GMV ($)" in h:
                return ExcelFormat.ENGLISH_USD
        return ExcelFormat.ENGLISH_USD

    raise UnknownFormatError(headers_str)


def _parse_decimal(value: object) -> Decimal:
    """将单元格值转换为 Decimal。"""
    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", ""))
    except InvalidOperation:
        return Decimal("0")


def _parse_int(value: object) -> int:
    """将单元格值转换为 int。"""
    if value is None or value == "":
        return 0
    try:
        return int(float(str(value).replace(",", "")))
    except (ValueError, TypeError):
        return 0


def read_excel_file(file_path: Path) -> list[VideoRecord]:
    """读取单个 Excel 文件，返回规范化的视频记录列表。

    如果文件格式不匹配（找不到表头），抛出 UnknownFormatError。
    """
    rows = _load_rows(file_path)
    if not rows:
        raise UnknownFormatError(f"[{file_path.name}] 文件无内容或无活动工作表")

    logger.debug("[%s] 共读取 %d 行原始数据", file_path.name, len(rows))

    # 动态查找表头行：扫描前 10 行，找到包含关键词的行
    header_idx: int | None = None
    for i, row in enumerate(rows[:10]):
        if _is_header_row(row):
            header_idx = i
            break

    if header_idx is None:
        if _is_empty_export(rows):
            logger.info("[%s] 识别为空导出文件，跳过", file_path.name)
            return []
        preview_lines = []
        for i, row in enumerate(rows[:5]):
            cells = [str(c) if c is not None else "(空)" for c in row[:5]]
            preview_lines.append(f"  第{i + 1}行: {' | '.join(cells)}")
        raise UnknownFormatError(
            f"[{file_path.name}] 在前10行中未找到表头"
            f"（需含'达人昵称'或'Creator name'），共{len(rows)}行。"
            f"前几行内容:\n" + "\n".join(preview_lines)
        )

    header_row = [str(cell).strip() if cell is not None else "" for cell in rows[header_idx]]
    logger.debug("[%s] 表头在第 %d 行: %s", file_path.name, header_idx + 1, header_row[:6])

    fmt = detect_format(header_row)
    if fmt == ExcelFormat.ENGLISH_GBP:
        currency = Currency.GBP
    elif fmt in {ExcelFormat.ENGLISH_JPY, ExcelFormat.CHINESE_JPY}:
        currency = Currency.JPY
    else:
        currency = Currency.USD

    col_map = _FORMAT_COLUMNS[fmt]

    # 建立列名到索引的映射
    col_indices: dict[str, int] = {}
    for field_name, col_header in col_map.items():
        for i, h in enumerate(header_row):
            if _matches_header(h, col_header):
                col_indices[field_name] = i
                break
        else:
            logger.warning(
                "[%s] 未找到列 '%s'，将使用默认值",
                file_path.name,
                _header_label(col_header),
            )

    if "creator_name" not in col_indices or "video_id" not in col_indices:
        raise UnknownFormatError(f"[{file_path.name}] 缺少必要列（creator_name 或 video_id）")

    creator_idx = col_indices["creator_name"]
    video_idx = col_indices["video_id"]
    product_idx = col_indices.get("product_name")
    video_gmv_idx = col_indices.get("video_gmv")
    gmv_idx = col_indices.get("attributed_gmv")
    orders_idx = col_indices.get("orders")
    items_idx = col_indices.get("items_sold")

    # 数据行 = 表头行之后的所有行
    data_rows = rows[header_idx + 1 :]
    records: list[VideoRecord] = []
    skipped = 0

    for row in data_rows:
        if not row or all(cell is None or cell == "" for cell in row):
            continue

        creator_name = str(row[creator_idx]).strip() if row[creator_idx] else ""
        video_id = str(row[video_idx]).strip() if row[video_idx] else ""

        if not creator_name or not video_id:
            skipped += 1
            continue

        product_name = (
            str(row[product_idx]).strip() if product_idx is not None and row[product_idx] else ""
        )
        video_gmv = (
            _parse_decimal(row[video_gmv_idx]) if video_gmv_idx is not None else Decimal("0")
        )
        attributed_gmv = _parse_decimal(row[gmv_idx]) if gmv_idx is not None else Decimal("0")
        orders = _parse_int(row[orders_idx]) if orders_idx is not None else 0
        items_sold = _parse_int(row[items_idx]) if items_idx is not None else 0

        records.append(
            VideoRecord(
                creator_name=creator_name,
                video_id=video_id,
                product_name=product_name,
                attributed_gmv=attributed_gmv,
                orders=orders,
                items_sold=items_sold,
                currency=currency,
                source_file=str(file_path),
                video_gmv=video_gmv,
            )
        )

    if skipped > 0:
        logger.debug("[%s] 跳过 %d 行（缺少达人/视频ID）", file_path.name, skipped)

    logger.info(
        "[%s] 格式=%s, 表头在第%d行, 读取 %d 条记录",
        file_path.name,
        fmt.value,
        header_idx + 1,
        len(records),
    )
    return records


def read_all_excel_files(input_dir: str | Path) -> list[VideoRecord]:
    """读取目录下所有 .xlsx 文件。

    任一文件格式不匹配将抛出异常中断执行。
    """
    dir_path = Path(input_dir)
    if not dir_path.exists():
        logger.warning("输入目录不存在: %s", dir_path)
        return []

    xlsx_files = sorted(dir_path.glob("*.xlsx"))
    if not xlsx_files:
        logger.warning("输入目录中没有 .xlsx 文件: %s", dir_path)
        return []

    logger.info("找到 %d 个 xlsx 文件", len(xlsx_files))

    all_records: list[VideoRecord] = []
    for f in xlsx_files:
        if f.name.startswith("~$"):
            logger.debug("跳过临时文件: %s", f.name)
            continue
        records = read_excel_file(f)
        all_records.extend(records)

    return all_records
